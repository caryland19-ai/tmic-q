#!/usr/bin/env python3
"""Generate a TMIC import workbook from a template and confirmed concept points.

The script intentionally keeps internal template IDs out of public output sheets.
It uses IDs only as parsing locators.
"""

from __future__ import annotations

import argparse
import json
import re
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


CHOICE_TYPES = {"单选", "多选", "编码（便于做均值分析）"}
TEXT_TYPES = {"填空"}
TARGET_IDS = {"Q3-1", "Q4-1", "Q6-1"}
NOT_CREDIBLE_ID = "Q6-1"
FOLLOWUP_ID_PREFIX = "Q6-2"
INTERNAL_ID_RE = re.compile(r"\b[A-Z]\d+(?:-\d+)?\b")
PLACEHOLDER_VALUES = {"XX", "XXX", "品牌1", "品牌2", "品牌3", "……", "..."}
GENERAL_LOGIC_KEYWORDS = (
    "基于项目需求",
    "需筛选",
    "筛选",
    "甄别",
    "目标人群",
    "购买品牌",
    "未选中",
    "建议",
    "针对",
)


@dataclass
class Question:
    internal_id: str
    text: str
    qtype: str
    page: int = 1
    options: list[str] = field(default_factory=list)
    option_codes: dict[str, str] = field(default_factory=dict)
    option_logic: dict[str, list[str]] = field(default_factory=dict)
    logic_notes: list[str] = field(default_factory=list)
    is_generated: bool = False


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"[ \t]+", " ", text)


def strip_internal_prefix(text: str) -> str:
    text = clean_text(text)
    text = re.sub(r"^\s*[A-Z]\d+(?:-\d+)?[\s.、:：-]+", "", text)
    return text.strip()


def public_type(qtype: str) -> str:
    if qtype == "单选" or qtype == "编码（便于做均值分析）":
        return "单选题"
    if qtype == "多选":
        return "多选题"
    if qtype == "填空":
        return "多行文本题"
    if qtype.endswith("题"):
        return qtype
    return qtype or "多行文本题"


def looks_like_question_type(value: str) -> bool:
    return value in CHOICE_TYPES or value in TEXT_TYPES or value.endswith("题")


def parse_template(path: Path, sheet_name: str | None) -> list[Question]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb["问卷模板"] if "问卷模板" in wb.sheetnames else wb.active
    questions: list[Question] = []
    current: Question | None = None
    current_page = 0

    for row in ws.iter_rows(values_only=True):
        cols = [clean_text(v) for v in row[:5]]
        section_marker, locator, text, qtype, logic = cols[0], cols[1], cols[2], cols[3], cols[4]

        if section_marker == "必答题":
            current_page += 1

        if looks_like_question_type(qtype):
            if current:
                questions.append(current)
            internal_id = locator or extract_internal_id(text) or f"ROW{len(questions) + 1}"
            current = Question(
                internal_id=internal_id,
                text=strip_internal_prefix(text),
                qtype=qtype,
                page=max(current_page, 1),
                logic_notes=[logic] if logic else [],
            )
            continue

        if current and text:
            current.options.append(text)
            if qtype:
                current.option_codes[text] = qtype
            if logic:
                current.option_logic.setdefault(text, []).append(logic)
                current.logic_notes.append(f"{text}: {logic}")
        elif current and logic:
            current.logic_notes.append(logic)

    if current:
        questions.append(current)
    return questions


def extract_internal_id(text: str) -> str:
    match = INTERNAL_ID_RE.search(text or "")
    return match.group(0) if match else ""


def replace_concept_points(questions: list[Question], points: list[str]) -> None:
    fixed_options = {
        "Q3-1": "以上描述都不满足我的需求",
        "Q4-1": "没有不太理解的词句",
        "Q6-1": "没有不太可信的描述",
    }
    for question in questions:
        if question.internal_id in TARGET_IDS:
            question.options = list(points) + [fixed_options[question.internal_id]]
            question.option_codes = {option: str(idx) for idx, option in enumerate(question.options, start=1)}
            question.option_logic = {fixed_options[question.internal_id]: ["互斥"]}

    not_credible_index = next((i for i, q in enumerate(questions) if q.internal_id == NOT_CREDIBLE_ID), None)
    if not_credible_index is None:
        return

    insert_at = not_credible_index + 1
    generated = [
        Question(
            internal_id=f"{FOLLOWUP_ID_PREFIX}-{idx}",
            text=f"请问您不相信“{point}”词句的原因是？",
            qtype="填空",
            page=questions[not_credible_index].page,
            is_generated=True,
        )
        for idx, point in enumerate(points, start=1)
    ]

    questions[:] = [
        q
        for q in questions
        if not (q.internal_id.startswith(f"{FOLLOWUP_ID_PREFIX}-") or q.internal_id == FOLLOWUP_ID_PREFIX)
    ]
    not_credible_index = next((i for i, q in enumerate(questions) if q.internal_id == NOT_CREDIBLE_ID), None)
    if not_credible_index is not None:
        questions[not_credible_index + 1 : not_credible_index + 1] = generated


def final_sequence(questions: list[Question]) -> dict[str, int]:
    return {q.internal_id: idx for idx, q in enumerate(questions, start=1)}


def question_label(question: Question, seq: dict[str, int]) -> str:
    return f"Q{seq[question.internal_id]}.{question.text}"


def option_texts_for_code_expr(question: Question, code_expr: str) -> list[str]:
    code_expr = clean_text(code_expr)
    wanted_positions: list[int] = []
    if "-" in code_expr:
        start, end = code_expr.split("-", 1)
        if start.isdigit() and end.isdigit():
            wanted_positions = list(range(int(start), int(end) + 1))
    elif code_expr:
        wanted_positions = [int(item.strip()) for item in re.split(r"[,，/、]", code_expr) if item.strip().isdigit()]

    if not wanted_positions:
        return []
    return [option for idx, option in enumerate(question.options, start=1) if idx in wanted_positions]


def render_question_logic_note(note: str, questions: list[Question], current: Question, seq: dict[str, int]) -> str:
    note = clean_text(note)
    questions_by_id = {q.internal_id: q for q in questions}
    match = re.search(r"(?:针对)?([A-Z]\d+(?:-\d+)?)中选中([0-9]+(?:-[0-9]+)?)(?:提问)?", note)
    if match:
        source_id, code_expr = match.group(1), match.group(2)
        source = next((q for q in questions if q.internal_id == source_id), None)
        if source:
            options = option_texts_for_code_expr(source, code_expr)
            if options:
                return f"如果「{question_label(source, seq)}」选择「{'、'.join(options)}」则显示本题，否则不显示。"
            return f"如果「{question_label(source, seq)}」选中代码 {code_expr} 则显示本题，否则不显示。"
    return INTERNAL_ID_RE.sub(lambda m: f"Q{seq[questions_by_id[m.group(0)].internal_id]}" if m.group(0) in questions_by_id else m.group(0), note)


def render_question_logic(question: Question, questions: list[Question]) -> str:
    seq = final_sequence(questions)
    if question.is_generated and question.internal_id.startswith(FOLLOWUP_ID_PREFIX):
        control = next((q for q in questions if q.internal_id == NOT_CREDIBLE_ID), None)
        point_match = re.search(r"请问您不相信“(.+)”词句的原因是？", question.text)
        if control and point_match:
            return f"如果「{question_label(control, seq)}」选择「{point_match.group(1)}」则显示本题，否则不显示。"

    rendered: list[str] = []
    for note in question.logic_notes:
        if ":" in note:
            continue
        rendered.append(render_question_logic_note(note, questions, question, seq))
    return "\n".join(item for item in rendered if item)


def render_option_logic(question: Question, option: str) -> str:
    return "\n".join(question.option_logic.get(option, []))


def import_rows(questions: list[Question], include_logic: bool = False) -> list[list[str]]:
    rows: list[list[str]] = []
    if include_logic:
        rows.append(["题号", "问卷导入内容", "逻辑"])
        seq = final_sequence(questions)
    for q in questions:
        question_row = [f"[{public_type(q.qtype)}]{q.text}"]
        if include_logic:
            question_row = [f"Q{seq[q.internal_id]}", question_row[0], render_question_logic(q, questions)]
        rows.append(question_row)
        if q.qtype in CHOICE_TYPES or public_type(q.qtype) in {"单选题", "多选题"}:
            for option in q.options:
                option_row = [option]
                if include_logic:
                    option_row = ["", option, render_option_logic(q, option)]
                rows.append(option_row)
        rows.append(["", "", ""] if include_logic else [""])
    if rows and rows[-1] in ([""], ["", ""], ["", "", ""]):
        rows.pop()
    return rows


def answer_logic_rows(questions: list[Question], points: list[str]) -> list[list[str]]:
    seq = final_sequence(questions)
    rows = [["整卷答题逻辑"]]
    control = next((q for q in questions if q.internal_id == NOT_CREDIBLE_ID), None)
    if control:
        for point in points:
            target = next((q for q in questions if q.text == f"请问您不相信“{point}”词句的原因是？"), None)
            if target:
                rows.append([
                    f"如果「Q{seq[control.internal_id]}.{control.text}」选择「{point}」则显示「Q{seq[target.internal_id]}.{target.text}」，否则不显示。"
                ])
    if len(rows) == 1:
        rows.append(["未从表格中读取到对应逻辑。"])
    return rows


def screener_logic_rows(questions: list[Question]) -> list[list[str]]:
    seq = final_sequence(questions)
    rows = [["整卷甄别逻辑"]]
    for q in questions:
        for note in q.logic_notes:
            if "终止" in note and ":" in note:
                option = note.split(":", 1)[0]
                rows.append([f"如果「Q{seq[q.internal_id]}.{q.text}」选择「{option}」则问卷结束，否则继续。"])
    if len(rows) == 1:
        rows.append(["未从表格中读取到对应逻辑。"])
    return rows


def chinese_page_label(page: int) -> str:
    labels = {
        1: "第一页",
        2: "第二页",
        3: "第三页",
        4: "第四页",
        5: "第五页",
        6: "第六页",
        7: "第七页",
        8: "第八页",
        9: "第九页",
        10: "第十页",
    }
    return labels.get(page, f"第{page}页")


def page_guidance_rows(questions: list[Question]) -> list[list[str]]:
    seq = final_sequence(questions)
    groups: dict[int, list[int]] = {}
    for q in questions:
        groups.setdefault(q.page, []).append(seq[q.internal_id])

    rows = [["分页指导"]]
    if not groups:
        rows.append(["第一页：Q1-Q0"])
        return rows

    for page in sorted(groups):
        numbers = groups[page]
        start, end = min(numbers), max(numbers)
        rows.append([f"{chinese_page_label(page)}：Q{start}-Q{end}"])
    return rows


def validation_rows(questions: list[Question], points: list[str]) -> tuple[list[list[str]], dict[int, str]]:
    rows = [["等级", "检查项", "意见"]]
    row_styles: dict[int, str] = {}
    public_text = "\n".join([q.text for q in questions] + [opt for q in questions for opt in q.options])
    seq = final_sequence(questions)

    def add(level: str, item: str, message: str) -> None:
        rows.append([level, item, message])
        if level == "重大问题":
            row_styles[len(rows)] = "major"
        elif level == "一般问题":
            row_styles[len(rows)] = "general"

    leaked = INTERNAL_ID_RE.findall(public_text)
    if leaked:
        add("重大问题", "内部编号暴露", "检测到疑似内部编号或源表编号，请确认是否来自题文原文；不得暴露模板内部 ID。")
    else:
        add("无问题", "内部编号暴露", "未检测到疑似模板内部编号或源表编号。")

    for target_id in TARGET_IDS:
        q = next((item for item in questions if item.internal_id == target_id), None)
        if not q:
            add("重大问题", "概念信息点替换", "未找到概念替换目标题，无法完成严格复刻。")
            continue
        actual = [opt for opt in q.options if opt in points]
        if len(actual) != len(points):
            add("重大问题", f"Q{seq[q.internal_id]} 概念信息点替换", f"概念选项数量为 {len(actual)}，确认信息点数量为 {len(points)}，不一致。")
        else:
            add("无问题", f"Q{seq[q.internal_id]} 概念信息点替换", f"已按确认信息点生成 {len(points)} 个选项。")

    followups = [q for q in questions if q.is_generated and q.internal_id.startswith(FOLLOWUP_ID_PREFIX)]
    if len(followups) != len(points):
        add("重大问题", "不可信原因追问题", f"不可信原因追问题数量为 {len(followups)}，确认信息点数量为 {len(points)}，不一致。")
    else:
        add("无问题", "不可信原因追问题", f"已按确认信息点生成 {len(points)} 道不可信原因追问题。")

    answer_logic = answer_logic_rows(questions, points)
    generated_answer_rules = max(len(answer_logic) - 1, 0) if answer_logic[1:] != [["未从表格中读取到对应逻辑。"]] else 0
    if generated_answer_rules != len(points):
        add("重大问题", "整卷答题逻辑", f"不可信原因显示逻辑数量为 {generated_answer_rules}，确认信息点数量为 {len(points)}，不一致。")
    else:
        add("无问题", "整卷答题逻辑", f"已生成 {generated_answer_rules} 条不可信原因显示逻辑。")

    explicit_terminations = [
        (q, note.split(":", 1)[0])
        for q in questions
        for note in q.logic_notes
        if "终止" in note and ":" in note
    ]
    if explicit_terminations:
        add("无问题", "整卷甄别逻辑", f"已读取到 {len(explicit_terminations)} 条明确终止规则。")
    else:
        add("一般问题", "整卷甄别逻辑", "未读取到明确终止规则；如问卷需要筛选终止，请人工确认。")

    for q in questions:
        qno = f"Q{seq[q.internal_id]}"
        placeholder_options = [opt for opt in q.options if opt in PLACEHOLDER_VALUES or "XX" in opt]
        if "XX" in q.text:
            add("一般问题", f"{qno} 题文占位", f"题文仍包含 `XX` 占位：{q.text}。如需上线或导入后展示，请替换为本项目真实品类/对象。")
        if placeholder_options:
            joined = "、".join(placeholder_options)
            add("一般问题", f"{qno} 选项占位", f"选项仍包含模板占位：{joined}。如这些选项参与筛选或逻辑判断，需替换为真实选项并同步逻辑。")

        for note in q.logic_notes:
            if not note:
                continue
            option = ""
            detail = note
            if ":" in note:
                option, detail = note.split(":", 1)
            if note == "选项随机":
                add("一般问题", f"{qno} 选项随机", "模板要求选项随机；导入后需确认 TMIC 后台是否已设置随机。")
            elif detail == "互斥":
                option_text = f"选项「{option}」" if option else "相关选项"
                add("一般问题", f"{qno} 互斥设置", f"{option_text}模板要求互斥；导入后需确认 TMIC 后台是否已设置互斥。")
            elif "终止" in note and ":" not in note:
                add("一般问题", f"{qno} 终止条件不完整", f"模板写有终止/筛选逻辑，但没有明确具体选项或条件：{note}。需要人工补充，否则不生成具体甄别逻辑。")
            elif any(keyword in detail for keyword in GENERAL_LOGIC_KEYWORDS):
                option_text = f"选项「{option}」" if option else "本题"
                add("一般问题", f"{qno} 逻辑需确认", f"{option_text}存在需要按项目定义确认的逻辑说明：{detail}。")

    return rows, row_styles


def write_sheet(ws, rows: list[list[str]], row_styles: dict[int, str] | set[int] | None = None) -> None:
    if isinstance(row_styles, set):
        row_styles = {idx: "major" for idx in row_styles}
    row_styles = row_styles or {}
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    red_font = Font(color="C00000")
    orange_fill = PatternFill("solid", fgColor="FCE4D6")
    orange_font = Font(color="ED7D31")
    black_font = Font(color="000000")
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
            if r_idx == 1:
                cell.font = Font(bold=True)
            elif row_styles.get(r_idx) == "major":
                cell.fill = red_fill
                cell.font = red_font
            elif row_styles.get(r_idx) == "general":
                cell.fill = orange_fill
                cell.font = orange_font
            else:
                cell.font = black_font
    for col in range(1, max((len(r) for r in rows), default=1) + 1):
        if rows and rows[0][:3] == ["题号", "问卷导入内容", "逻辑"]:
            width = {1: 12, 2: 80, 3: 70}.get(col, 40)
        else:
            width = 80 if col == 1 else 40
        ws.column_dimensions[get_column_letter(col)].width = width


def style_complete_import_sheet(ws) -> None:
    logic_font_color = "808080"
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row_idx, 3)
        if cell.value in (None, ""):
            continue
        cell.font = copy(cell.font)
        cell.font = Font(
            name=cell.font.name,
            sz=cell.font.sz,
            b=cell.font.bold,
            i=cell.font.italic,
            color=logic_font_color,
            underline=cell.font.underline,
        )


def build_workbook(questions: list[Question], points: list[str], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "完整TMIC问卷导入表格"
    write_sheet(ws, import_rows(questions, include_logic=True))
    style_complete_import_sheet(ws)

    validation, validation_styles = validation_rows(questions, points)
    write_sheet(wb.create_sheet("逻辑校验结果"), validation, validation_styles)
    write_sheet(wb.create_sheet("整卷答题逻辑"), answer_logic_rows(questions, points))
    write_sheet(wb.create_sheet("整卷甄别逻辑"), screener_logic_rows(questions))
    write_sheet(wb.create_sheet("分页指导"), page_guidance_rows(questions))
    write_sheet(wb.create_sheet("信息点确认"), [["已确认信息点"]] + [[p] for p in points])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def build_upload_workbook(questions: list[Question], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TMIC上传专用表格"
    write_sheet(ws, import_rows(questions, include_logic=False))

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def load_points(path_or_json: str) -> list[str]:
    candidate = Path(path_or_json)
    if candidate.exists():
        text = candidate.read_text(encoding="utf-8")
    else:
        text = path_or_json
    data = json.loads(text)
    if isinstance(data, list):
        return [clean_text(item) for item in data if clean_text(item)]
    if isinstance(data, dict) and isinstance(data.get("points"), list):
        return [clean_text(item) for item in data["points"] if clean_text(item)]
    raise ValueError("Concept points must be a JSON list or an object with a points list.")


def main() -> None:
    skill_root = Path(__file__).resolve().parents[1]
    default_template = skill_root / "assets" / "TMIC模版.xlsx"
    default_import_format = skill_root / "assets" / "问卷格式.xlsx"

    parser = argparse.ArgumentParser()
    parser.add_argument("--template", default=default_template, type=Path)
    parser.add_argument("--import-format", default=default_import_format, type=Path)
    parser.add_argument("--points", required=True, help="JSON list or path to JSON file.")
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--upload-output", default=None, type=Path)
    parser.add_argument("--sheet-name", default=None)
    args = parser.parse_args()

    if not args.template.exists():
        raise FileNotFoundError(f"Missing TMIC template workbook: {args.template}")
    if not args.import_format.exists():
        raise FileNotFoundError(f"Missing TMIC import-format workbook: {args.import_format}")
    load_workbook(args.import_format, read_only=True, data_only=True).close()

    points = load_points(args.points)
    if not points:
        raise ValueError("No confirmed concept points provided.")
    questions = parse_template(args.template, args.sheet_name)
    if not questions:
        raise ValueError("No questions parsed from template workbook.")
    replace_concept_points(questions, points)
    build_workbook(questions, points, args.output)
    if args.upload_output:
        build_upload_workbook(questions, args.upload_output)
    print(args.output)
    if args.upload_output:
        print(args.upload_output)


if __name__ == "__main__":
    main()
