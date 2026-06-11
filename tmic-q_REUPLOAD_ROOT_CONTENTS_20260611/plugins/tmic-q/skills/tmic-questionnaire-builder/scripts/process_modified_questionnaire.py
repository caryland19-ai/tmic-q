#!/usr/bin/env python3
"""Rebuild TMIC outputs from a user-modified complete questionnaire workbook."""

from __future__ import annotations

import argparse
import re
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


QUESTION_PREFIX_RE = re.compile(r"^\[(?P<qtype>[^\]]+)\](?P<text>.*)$")
QNO_RE = re.compile(r"^Q?(?P<num>\d+)$", re.IGNORECASE)
Q_REF_RE = re.compile(r"Q(?P<num>\d+)")
QUOTED_NUM_REF_RE = re.compile(r"「(?P<num>\d+)\.")
QUOTED_QUESTION_REF_RE = re.compile(r"「Q?(?P<num>\d+)\.(?P<text>[^」]*)」")
PLACEHOLDER_VALUES = ("XX", "XXX", "品牌1", "品牌2", "品牌3", "……", "...")
CHOICE_TYPES = {"单选题", "多选题", "单选", "多选", "编码（便于做均值分析）"}
LOGIC_GRAY = "808080"


@dataclass
class Option:
    text: str
    logic: str = ""


@dataclass
class Question:
    old_qno: str
    qtype: str
    text: str
    heading: str
    logic: str = ""
    options: list[Option] = field(default_factory=list)
    new_index: int = 0

    @property
    def new_qno(self) -> str:
        return f"Q{self.new_index}"

    @property
    def new_label(self) -> str:
        return f"{self.new_qno}.{self.text}"


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def parse_qno(value: str) -> str:
    match = QNO_RE.match(cell_text(value))
    return match.group("num") if match else ""


def parse_question_heading(content: str) -> tuple[str, str]:
    match = QUESTION_PREFIX_RE.match(content)
    if not match:
        return "", content
    return cell_text(match.group("qtype")), cell_text(match.group("text"))


def is_question_row(qno: str, content: str) -> bool:
    return bool(parse_qno(qno)) or bool(QUESTION_PREFIX_RE.match(content))


def find_import_sheet(wb) -> Any:
    if "完整TMIC问卷导入表格" in wb.sheetnames:
        return wb["完整TMIC问卷导入表格"]
    return wb.active


def find_import_columns(ws) -> tuple[int, int, int, int]:
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        values = [cell_text(ws.cell(row_idx, col).value) for col in range(1, min(ws.max_column, 8) + 1)]
        if "问卷导入内容" in values:
            content_col = values.index("问卷导入内容") + 1
            qno_col = values.index("题号") + 1 if "题号" in values else 0
            logic_col = values.index("逻辑") + 1 if "逻辑" in values else 0
            return row_idx, qno_col, content_col, logic_col
    return 0, 0, 1, 0


def read_questions(path: Path) -> tuple[list[Question], list[str]]:
    wb = load_workbook(path, data_only=True)
    ws = find_import_sheet(wb)
    header_row, qno_col, content_col, logic_col = find_import_columns(ws)
    start_row = header_row + 1 if header_row else 1
    warnings: list[str] = []
    questions: list[Question] = []
    current: Question | None = None

    if not header_row:
        warnings.append("未识别到 `问卷导入内容` 表头；已按第一列为导入内容尝试解析。")
    if not logic_col:
        warnings.append("未识别到 `逻辑` 列；将只能基于题目和选项输出上传专用表格。")

    for row_idx in range(start_row, ws.max_row + 1):
        qno = cell_text(ws.cell(row_idx, qno_col).value) if qno_col else ""
        content = cell_text(ws.cell(row_idx, content_col).value)
        logic = cell_text(ws.cell(row_idx, logic_col).value) if logic_col else ""
        if not content and not qno and not logic:
            continue

        if content and is_question_row(qno, content):
            qtype, text = parse_question_heading(content)
            if not qtype:
                qtype = "多行文本题"
            current = Question(
                old_qno=parse_qno(qno),
                qtype=qtype,
                text=text,
                heading=content if QUESTION_PREFIX_RE.match(content) else f"[{qtype}]{text}",
                logic=logic,
            )
            questions.append(current)
            continue

        if current and content:
            current.options.append(Option(content, logic))
        elif logic:
            warnings.append(f"第 {row_idx} 行只有逻辑没有问卷内容，已忽略：{logic}")

    for idx, question in enumerate(questions, start=1):
        question.new_index = idx
    return questions, warnings


def old_to_new_map(questions: list[Question]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for question in questions:
        if question.old_qno:
            mapping[question.old_qno] = str(question.new_index)
    return mapping


def adjust_logic_text(text: str, mapping: dict[str, str]) -> str:
    if not text:
        return ""

    def replace_q_ref(match: re.Match[str]) -> str:
        old = match.group("num")
        return f"Q{mapping.get(old, old)}"

    def replace_quoted_ref(match: re.Match[str]) -> str:
        old = match.group("num")
        return f"「{mapping.get(old, old)}."

    text = Q_REF_RE.sub(replace_q_ref, text)
    return QUOTED_NUM_REF_RE.sub(replace_quoted_ref, text)


def all_adjusted_logic(questions: list[Question]) -> list[tuple[Question, str, str]]:
    mapping = old_to_new_map(questions)
    result: list[tuple[Question, str, str]] = []
    for question in questions:
        if question.logic:
            result.append((question, "", adjust_logic_text(question.logic, mapping)))
        for option in question.options:
            if option.logic:
                result.append((question, option.text, adjust_logic_text(option.logic, mapping)))
    return result


def referenced_numbers(text: str) -> set[int]:
    numbers = {int(match.group("num")) for match in Q_REF_RE.finditer(text)}
    numbers.update(int(match.group("num")) for match in QUOTED_NUM_REF_RE.finditer(text))
    return numbers


def quoted_question_refs(text: str) -> list[tuple[int, str]]:
    return [(int(match.group("num")), cell_text(match.group("text"))) for match in QUOTED_QUESTION_REF_RE.finditer(text)]


def split_option_condition(condition: str) -> list[str]:
    return [item.strip() for item in re.split(r"[、,，/]", condition) if item.strip()]


def normalize_option_text(text: str) -> str:
    return re.sub(r"[\s、,，/]+", "", text)


def normalize_question_text(text: str) -> str:
    return re.sub(r"[\s，,。？?：:；;、]+", "", text)


def condition_matches_options(condition: str, options: set[str]) -> bool:
    if condition in options:
        return True
    normalized_condition = normalize_option_text(condition)
    normalized_options = {normalize_option_text(option) for option in options}
    if normalized_condition in normalized_options:
        return True

    remaining = normalized_condition
    for option in sorted(normalized_options, key=len, reverse=True):
        if option and option in remaining:
            remaining = remaining.replace(option, "", 1)
    return remaining == ""


def validation_rows(questions: list[Question], warnings: list[str], base_questions: list[Question] | None) -> tuple[list[list[str]], dict[int, str]]:
    rows = [["等级", "检查项", "意见"]]
    styles: dict[int, str] = {}

    def add(level: str, item: str, message: str) -> None:
        rows.append([level, item, message])
        if level == "重大问题":
            styles[len(rows)] = "major"
        elif level == "一般问题":
            styles[len(rows)] = "general"

    if questions:
        add("无问题", "修改版问卷读取", f"已读取到 {len(questions)} 道最终题目，并按当前顺序重算题号。")
    else:
        add("重大问题", "修改版问卷读取", "未读取到任何题目，无法生成上传专用表格。")

    for warning in warnings:
        add("一般问题", "输入格式提醒", warning)

    if base_questions is None:
        add("一般问题", "变更识别范围", "未提供上一版完整问卷；本次仅基于修改版当前内容校验逻辑引用，无法逐项判断新增、删除、改写来源。")
    else:
        base_texts = [q.text for q in base_questions]
        current_texts = [q.text for q in questions]
        deleted = [text for text in base_texts if text not in current_texts]
        added = [text for text in current_texts if text not in base_texts]
        moved = sum(1 for text in current_texts if text in base_texts and current_texts.index(text) != base_texts.index(text))
        if deleted or added or moved:
            add("一般问题", "变更识别", f"识别到新增题 {len(added)} 道、删除题 {len(deleted)} 道、位置变化题 {moved} 道；已按修改版当前顺序重算题号和逻辑引用。")
        else:
            add("无问题", "变更识别", "未识别到题目新增、删除或移动。")

    question_by_num = {q.new_index: q for q in questions}
    logic_entries = all_adjusted_logic(questions)
    stale_refs: list[str] = []
    backward_display: list[str] = []
    missing_options: list[str] = []

    for source_question, option_text, logic in logic_entries:
        for number in sorted(referenced_numbers(logic)):
            if number not in question_by_num:
                stale_refs.append(f"{source_question.new_qno} 引用了不存在的 Q{number}：{logic}")

        for number, ref_text in quoted_question_refs(logic):
            target = question_by_num.get(number)
            if not target or not ref_text:
                continue
            if normalize_question_text(ref_text) != normalize_question_text(target.text):
                stale_refs.append(
                    f"{source_question.new_qno} 的逻辑引用 Q{number}.{ref_text}，但当前 Q{number} 是「{target.text}」；可能是删除/移动后逻辑未同步。"
                )

        source_match = re.search(r"如果「Q?(?P<num>\d+)\.[^」]*」选择「(?P<option>[^」]+)」", logic)
        if source_match:
            source_num = int(source_match.group("num"))
            source = question_by_num.get(source_num)
            if source:
                options = {opt.text for opt in source.options}
                condition = source_match.group("option")
                if condition_matches_options(condition, options):
                    missing = []
                else:
                    missing = [item for item in split_option_condition(condition) if item not in options]
                if missing and options:
                    missing_options.append(f"Q{source_num} 缺少逻辑条件选项：{'、'.join(missing)}")

        target_match = re.search(r"显示「Q?(?P<num>\d+)\.[^」]*」", logic)
        if source_match and target_match:
            source_num = int(source_match.group("num"))
            target_num = int(target_match.group("num"))
            if source_num in question_by_num and target_num in question_by_num and target_num <= source_num:
                backward_display.append(f"Q{source_num} 显示 Q{target_num}，目标题未位于控制题之后：{logic}")

    if stale_refs:
        add("重大问题", "失效题目引用", "；".join(stale_refs[:8]))
    else:
        add("无问题", "失效题目引用", "未检测到指向不存在题目的逻辑引用。")

    if missing_options:
        add("重大问题", "失效选项引用", "；".join(missing_options[:8]))
    else:
        add("无问题", "失效选项引用", "未检测到指向不存在选项的显示逻辑。")

    if backward_display:
        add("重大问题", "显示题顺序", "；".join(backward_display[:8]))
    else:
        add("无问题", "显示题顺序", "未检测到显示目标位于控制题之前的情况。")

    if not any("显示" in logic for _, _, logic in logic_entries):
        add("一般问题", "整卷答题逻辑", "未读取到显示逻辑；如修改后有条件显示题，请在完整问卷的逻辑列补充后重新处理。")
    else:
        add("无问题", "整卷答题逻辑", "已从修改版逻辑列读取并重算显示逻辑。")

    if not any("终止" in logic or "问卷结束" in logic for _, _, logic in logic_entries):
        add("一般问题", "整卷甄别逻辑", "未读取到终止/问卷结束逻辑；如修改后有筛选终止条件，请人工确认。")
    else:
        add("无问题", "整卷甄别逻辑", "已从修改版逻辑列读取并重算甄别逻辑。")

    for question in questions:
        texts = [question.text] + [option.text for option in question.options]
        if any(token and token in text for token in PLACEHOLDER_VALUES for text in texts):
            add("一般问题", f"{question.new_qno} 模板占位", "题文或选项仍包含模板占位；如需上线，请替换真实内容并同步逻辑。")
        if question.logic and any(keyword in question.logic for keyword in ("新增", "人工确认", "基于项目需求", "筛选", "甄别")):
            add("一般问题", f"{question.new_qno} 逻辑需确认", f"本题逻辑列包含需人工确认内容：{question.logic}")

    return rows, styles


def page_guidance_rows(questions: list[Question], source_path: Path) -> list[list[str]]:
    wb = load_workbook(source_path, data_only=True)
    if "分页指导" not in wb.sheetnames:
        if not questions:
            return [["分页指导"], ["第一页：Q1-Q0"]]
        return [["分页指导"], [f"第一页：Q1-Q{len(questions)}"]]

    ws = wb["分页指导"]
    mapping = old_to_new_map(questions)
    rows = [["分页指导"]]
    used_new_numbers: set[int] = set()

    def mapped_page_numbers(text: str) -> list[int]:
        old_numbers: set[int] = set()
        for start, end in re.findall(r"Q(\d+)\s*-\s*Q(\d+)", text):
            old_numbers.update(range(int(start), int(end) + 1))
        old_numbers.update(int(num) for num in re.findall(r"Q(\d+)", text))
        return sorted(int(mapping[str(num)]) for num in old_numbers if str(num) in mapping)

    for row in ws.iter_rows(values_only=True):
        text = cell_text(row[0] if row else "")
        if not text or text == "分页指导":
            continue
        label = text.split("：", 1)[0] if "：" in text else text.split(":", 1)[0]
        nums = mapped_page_numbers(text)
        if nums:
            used_new_numbers.update(nums)
            rows.append([f"{label}：Q{min(nums)}-Q{max(nums)}"])

    missing = [q.new_index for q in questions if q.new_index not in used_new_numbers]
    if missing:
        label = f"第{len(rows)}页" if len(rows) > 1 else "第一页"
        rows.append([f"{label}：Q{min(missing)}-Q{max(missing)}"])
    if len(rows) == 1:
        rows.append([f"第一页：Q1-Q{len(questions)}"])
    return rows


def answer_logic_rows(questions: list[Question]) -> list[list[str]]:
    rows = [["整卷答题逻辑"]]
    for question, _, logic in all_adjusted_logic(questions):
        if "显示" in logic and "终止" not in logic and "问卷结束" not in logic:
            rows.append([logic.replace("显示本题", f"显示「{question.new_label}」")])
    if len(rows) == 1:
        rows.append(["未从表格中读取到对应逻辑。"])
    return rows


def screener_logic_rows(questions: list[Question]) -> list[list[str]]:
    rows = [["整卷甄别逻辑"]]
    for question, option_text, logic in all_adjusted_logic(questions):
        if "终止" in logic or "问卷结束" in logic:
            if option_text:
                rows.append([f"如果「{question.new_label}」选择「{option_text}」则问卷结束，否则继续。"])
            else:
                rows.append([logic])
    if len(rows) == 1:
        rows.append(["未从表格中读取到对应逻辑。"])
    return rows


def change_rows(base_questions: list[Question] | None, questions: list[Question]) -> list[list[str]]:
    rows = [["变更识别"]]
    if base_questions is None:
        rows.append(["未提供上一版完整问卷，无法逐项识别新增、删除、改写和移动。"])
        return rows

    base_by_text = {q.text: idx for idx, q in enumerate(base_questions, start=1)}
    current_by_text = {q.text: idx for idx, q in enumerate(questions, start=1)}
    for text, idx in current_by_text.items():
        if text not in base_by_text:
            rows.append([f"新增：Q{idx}.{text}"])
        elif base_by_text[text] != idx:
            rows.append([f"移动：原 Q{base_by_text[text]} -> Q{idx}.{text}"])
    for text, idx in base_by_text.items():
        if text not in current_by_text:
            rows.append([f"删除：原 Q{idx}.{text}"])
    if len(rows) == 1:
        rows.append(["未识别到题目新增、删除或移动。"])
    return rows


def write_sheet(ws, rows: list[list[str]], row_styles: dict[int, str] | None = None, gray_col: int | None = None) -> None:
    row_styles = row_styles or {}
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    red_font = Font(color="C00000")
    orange_fill = PatternFill("solid", fgColor="FCE4D6")
    orange_font = Font(color="ED7D31")
    black_font = Font(color="000000")
    gray_font = Font(color=LOGIC_GRAY)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c_idx, value)
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "top"
            cell.alignment = alignment
            if r_idx == 1:
                cell.font = Font(bold=True, color=LOGIC_GRAY if c_idx == gray_col else "000000")
            elif gray_col and c_idx == gray_col and value:
                cell.font = gray_font
            elif row_styles.get(r_idx) == "major":
                cell.fill = red_fill
                cell.font = red_font
            elif row_styles.get(r_idx) == "general":
                cell.fill = orange_fill
                cell.font = orange_font
            else:
                cell.font = black_font
    for col in range(1, max((len(row) for row in rows), default=1) + 1):
        width = {1: 12, 2: 80, 3: 70}.get(col, 80)
        ws.column_dimensions[get_column_letter(col)].width = width


def complete_import_rows(questions: list[Question]) -> list[list[str]]:
    rows = [["题号", "问卷导入内容", "逻辑"]]
    mapping = old_to_new_map(questions)
    for question in questions:
        rows.append([question.new_qno, question.heading, adjust_logic_text(question.logic, mapping)])
        for option in question.options:
            rows.append(["", option.text, adjust_logic_text(option.logic, mapping)])
        rows.append(["", "", ""])
    if rows and rows[-1] == ["", "", ""]:
        rows.pop()
    return rows


def upload_rows(questions: list[Question]) -> list[list[str]]:
    rows: list[list[str]] = []
    for question in questions:
        rows.append([question.heading])
        for option in question.options:
            rows.append([option.text])
        rows.append([""])
    if rows and rows[-1] == [""]:
        rows.pop()
    return rows


def save_outputs(input_path: Path, output_complete: Path, output_upload: Path, base_path: Path | None = None) -> None:
    questions, warnings = read_questions(input_path)
    base_questions = read_questions(base_path)[0] if base_path else None

    complete_wb = Workbook()
    ws = complete_wb.active
    ws.title = "完整TMIC问卷导入表格"
    write_sheet(ws, complete_import_rows(questions), gray_col=3)
    validation, styles = validation_rows(questions, warnings, base_questions)
    write_sheet(complete_wb.create_sheet("逻辑校验结果"), validation, styles)
    write_sheet(complete_wb.create_sheet("整卷答题逻辑"), answer_logic_rows(questions))
    write_sheet(complete_wb.create_sheet("整卷甄别逻辑"), screener_logic_rows(questions))
    write_sheet(complete_wb.create_sheet("分页指导"), page_guidance_rows(questions, input_path))
    write_sheet(complete_wb.create_sheet("变更识别"), change_rows(base_questions, questions))
    output_complete.parent.mkdir(parents=True, exist_ok=True)
    complete_wb.save(output_complete)

    upload_wb = Workbook()
    upload_ws = upload_wb.active
    upload_ws.title = "TMIC上传专用表格"
    write_sheet(upload_ws, upload_rows(questions))
    output_upload.parent.mkdir(parents=True, exist_ok=True)
    upload_wb.save(output_upload)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path, help="User-modified complete TMIC questionnaire workbook.")
    parser.add_argument("--output-complete", required=True, type=Path, help="Adjusted complete workbook path.")
    parser.add_argument("--output-upload", required=True, type=Path, help="Upload-only workbook path.")
    parser.add_argument("--base-workbook", default=None, type=Path, help="Optional previous complete workbook for change detection.")
    args = parser.parse_args()

    if not args.input.exists():
        raise FileNotFoundError(f"Missing modified questionnaire workbook: {args.input}")
    if args.base_workbook and not args.base_workbook.exists():
        raise FileNotFoundError(f"Missing base questionnaire workbook: {args.base_workbook}")
    save_outputs(args.input, args.output_complete, args.output_upload, args.base_workbook)
    print(args.output_complete)
    print(args.output_upload)


if __name__ == "__main__":
    main()
